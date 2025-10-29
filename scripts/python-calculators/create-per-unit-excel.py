#!/usr/bin/env python3
"""
Create Excel workbook with per-unit pricing breakdown
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the per-unit pricing data
df = pd.read_csv('G:/Shared drives/Energen Ops/2-Sales & Marketing/1-Sales/Bids/LNBL-Whole-Facility-RFP/per-unit-pricing.csv')

# Add the missing 3 generators (same pricing as matching kW)
missing_gens = [
    {"assetId": "02 EG 068", "manufacturer": "Cummins", "kwRating": 300, "serviceA": 6933.92, "serviceB": 1711.8, "serviceD": 93.1, "serviceE": 3680, "totalAnnual": 12418.82},
    {"assetId": "30 EG 114", "manufacturer": "Kohler", "kwRating": 350, "serviceA": 6933.92, "serviceB": 1711.8, "serviceD": 93.1, "serviceE": 3680, "totalAnnual": 12418.82},
    {"assetId": "33U EG 113", "manufacturer": "Cummins", "kwRating": 350, "serviceA": 6933.92, "serviceB": 1711.8, "serviceD": 93.1, "serviceE": 3680, "totalAnnual": 12418.82}
]

df_missing = pd.DataFrame(missing_gens)
df = pd.concat([df_missing, df], ignore_index=True)

# Sort by asset ID
df = df.sort_values('assetId').reset_index(drop=True)

# Calculate totals
total_annual = df['totalAnnual'].sum()
total_service_a = df['serviceA'].sum()
total_service_b = df['serviceB'].sum()
total_service_d = df['serviceD'].sum()
total_service_e = df['serviceE'].sum()

print(f"Total generators: {len(df)}")
print(f"Total annual cost: ${total_annual:,.2f}")

# Create Excel workbook
wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
currency_format = '$#,##0.00'
total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
total_font = Font(bold=True, size=11)

# Sheet 1: Per-Unit Pricing
ws = wb.create_sheet("PER-UNIT PRICING", 0)

# Set column widths
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 16

# Title
ws.merge_cells('A1:H1')
ws['A1'] = "LNBL RFP ANR-6-2025 - PER-UNIT PRICING"
ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws['A1'].fill = header_fill
ws['A1'].alignment = Alignment(horizontal='center')

# Headers
headers = ['Asset ID', 'Manufacturer', 'kW', 'Service A', 'Service B', 'Service D', 'Service E', 'Total Annual']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(3, col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Data rows
row = 4
for idx, gen in df.iterrows():
    ws[f'A{row}'] = gen['assetId']
    ws[f'B{row}'] = gen['manufacturer']
    ws[f'C{row}'] = int(gen['kwRating'])
    ws[f'D{row}'] = gen['serviceA']
    ws[f'E{row}'] = gen['serviceB']
    ws[f'F{row}'] = gen['serviceD']
    ws[f'G{row}'] = gen['serviceE']
    ws[f'H{row}'] = gen['totalAnnual']

    # Format currency columns
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{row}'].number_format = currency_format

    row += 1

# Total row
ws[f'A{row}'] = "TOTAL"
ws[f'D{row}'] = total_service_a
ws[f'E{row}'] = total_service_b
ws[f'F{row}'] = total_service_d
ws[f'G{row}'] = total_service_e
ws[f'H{row}'] = total_annual

ws[f'A{row}'].font = total_font
for col in ['D', 'E', 'F', 'G', 'H']:
    ws[f'{col}{row}'].font = total_font
    ws[f'{col}{row}'].fill = total_fill
    ws[f'{col}{row}'].number_format = currency_format

# Sheet 2: Summary by kW Range
ws_summary = wb.create_sheet("SUMMARY BY kW RANGE", 1)

# Group by kW range
def get_kw_range(kw):
    if kw <= 30:
        return "15-30 kW"
    elif kw <= 150:
        return "35-150 kW"
    elif kw <= 250:
        return "151-250 kW"
    elif kw <= 400:
        return "251-400 kW"
    elif kw <= 500:
        return "401-500 kW"
    elif kw <= 670:
        return "501-670 kW"
    elif kw <= 1050:
        return "671-1050 kW"
    elif kw <= 1500:
        return "1051-1500 kW"
    elif kw <= 2050:
        return "1501-2050 kW"
    else:
        return "2050+ kW"

df['kwRange'] = df['kwRating'].apply(get_kw_range)
summary = df.groupby('kwRange').agg({
    'assetId': 'count',
    'totalAnnual': 'sum'
}).rename(columns={'assetId': 'count', 'totalAnnual': 'totalCost'})

ws_summary.column_dimensions['A'].width = 20
ws_summary.column_dimensions['B'].width = 15
ws_summary.column_dimensions['C'].width = 20

ws_summary['A1'] = "kW Range"
ws_summary['B1'] = "Unit Count"
ws_summary['C1'] = "Total Annual Cost"

for cell in ['A1', 'B1', 'C1']:
    ws_summary[cell].font = header_font
    ws_summary[cell].fill = header_fill

row = 2
for kw_range, data in summary.iterrows():
    ws_summary[f'A{row}'] = kw_range
    ws_summary[f'B{row}'] = int(data['count'])
    ws_summary[f'C{row}'] = data['totalCost']
    ws_summary[f'C{row}'].number_format = currency_format
    row += 1

# Save workbook
output_file = 'G:/Shared drives/Energen Ops/2-Sales & Marketing/1-Sales/Bids/LNBL-Whole-Facility-RFP/LNBL-Per-Unit-Pricing.xlsx'
wb.save(output_file)

print(f"\n{'='*70}")
print(f"Excel workbook created: LNBL-Per-Unit-Pricing.xlsx")
print(f"{'='*70}")
print(f"\nSheets:")
print(f"  1. PER-UNIT PRICING - All 34 generators with service breakdown")
print(f"  2. SUMMARY BY kW RANGE - Cost summary grouped by kW tier")
print(f"\nTotal Annual Cost: ${total_annual:,.2f}")
print(f"{'='*70}")
