#!/usr/bin/env python3
"""
Analyze LBNL Pricing Excel Template Structure
"""
import openpyxl
import sys

excel_path = r'G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\5-Offerors Pricing.xlsx'

print(f"Loading: {excel_path}\n")

wb = openpyxl.load_workbook(excel_path, data_only=False)
print(f"Sheets: {wb.sheetnames}\n")

ws = wb['Pricing']
print(f"=== PRICING SHEET STRUCTURE ===\n")

# Scan first 50 rows to understand structure
for row in range(1, 51):
    row_data = []
    for col in range(1, 12):  # First 11 columns (A-K)
        cell = ws.cell(row, col)
        val = cell.value
        
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col)
            # Show formulas if present
            if cell.data_type == 'f':
                row_data.append(f"{col_letter}: [FORMULA] {val}")
            else:
                row_data.append(f"{col_letter}: {val}")
    
    if row_data:
        print(f"Row {row:2d}: {' | '.join(row_data)}")

print("\n=== COLUMN HEADERS ===")
for col in range(1, 15):
    header = ws.cell(1, col).value
    if header:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col_letter}: {header}")

wb.close()
print("\n✅ Analysis complete")
