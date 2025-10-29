"""
Rebuild Service Lookup Tables with ALL pricing from default-settings.json
NO HARDCODED VALUES - everything from settings
"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Paths
excel_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-Formula-Driven-FINAL-LABELED-All-36-Units.xlsx"
settings_path = r"C:\ECalc\active\energen-calculator-v5.0\frontend\config\default-settings.json"

# Load calculator settings
with open(settings_path, 'r') as f:
    settings = json.load(f)

# Extract pricing values
LABOR_RATE = settings['laborRate']
MOBILIZATION_RATE = settings['mobilizationRate']
OIL_BASE = settings['oilPrice']
OIL_MARKUP = settings['oilMarkup']
OIL_PRICE = OIL_BASE * OIL_MARKUP
COOLANT_BASE = settings['coolantPrice']
COOLANT_MARKUP = settings['coolantMarkup']
COOLANT_PRICE = COOLANT_BASE * COOLANT_MARKUP
PARTS_MARKUP = settings['partsMarkup']

wb = load_workbook(excel_path)

# Delete and recreate sheet
if "Service Lookup Tables" in wb.sheetnames:
    del wb["Service Lookup Tables"]
ws = wb.create_sheet("Service Lookup Tables", 1)

# Styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, size=13, color="FFFFFF")
subtitle_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
subtitle_font = Font(bold=True, size=11, color="FFFFFF")
col_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
col_header_font = Font(bold=True, size=10, color="000000")
category_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
note_font = Font(bold=True, size=10, color="C65911")
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Size categories
size_categories = {
    "2-14": "Very Small Units",
    "15-30": "Small Units",
    "35-150": "Medium Units",
    "151-250": "Large Units",
    "251-400": "Very Large Units",
    "401-500": "Extra Large Units",
    "501-670": "Industrial Units",
    "671-1050": "Large Industrial",
    "1051-1500": "Very Large Industrial",
    "1501+": "Extra Large Industrial"
}

row = 1

# ============================================================================
# SERVICE A - COMPREHENSIVE INSPECTION
# ============================================================================
print("Building Service A...")
service_a = settings['serviceA']

ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE A - COMPREHENSIVE INSPECTION"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = f"Annual Preventative Maintenance - Frequency: {service_a['frequency']} times per year (Quarterly)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

# Column headers
headers = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Parts & Supplies\nCost\n(inspection items)",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.border = border
    cell.fill = col_header_fill
    cell.font = col_header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[chr(64+col)].width = [14, 18, 12, 12, 14, 14, 14][col-1]
ws.row_dimensions[row].height = 45
row += 1

# Data rows
for kw_range, values in service_a['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).fill = category_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=3).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=5, value=values['parts']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=5).number_format = '$#,##0.00'

    # Total formula using laborRate from settings
    total = (values['labor'] + values['mobilization']) * LABOR_RATE + values['parts']
    ws.cell(row=row, column=6, value=total).border = border
    ws.cell(row=row, column=6).fill = total_fill
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=6).number_format = '$#,##0.00'
    ws.cell(row=row, column=6).font = Font(bold=True, size=10)

    ws.cell(row=row, column=7, value="4x per year").border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=7).font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

# Pricing note
row += 1
ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = f"Pricing Details: Labor & mobilization at ${LABOR_RATE:.2f}/hr (standard commercial rate). Parts include inspection supplies per service tier."
cell.fill = note_fill
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 30
row += 3

# ============================================================================
# SERVICE B - OIL & FILTER SERVICE
# ============================================================================
print("Building Service B...")
service_b = settings['serviceB']

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE B - OIL & FILTER SERVICE"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Engine Oil & Filter Replacement - Frequency: {service_b.get('frequency', 'Annual')}"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

# Column headers
headers_b = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Oil Filter\nReplacement\nCost",
    "Engine Oil\nVolume Needed\n(gallons)",
    "Oil Cost\n@ $"+f"{OIL_PRICE:.2f}"+"/gallon\n(with markup)",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col, header in enumerate(headers_b, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.border = border
    cell.fill = col_header_fill
    cell.font = col_header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[chr(64+col)].width = [14, 18, 12, 12, 12, 12, 12, 14, 14][col-1]
ws.row_dimensions[row].height = 45
row += 1

# Data rows
for kw_range, values in service_b['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).fill = category_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=3).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=5, value=values['filterCost']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=5).number_format = '$#,##0.00'

    ws.cell(row=row, column=6, value=values['oilGallons']).border = border
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=6).number_format = '0.0 "gal"'

    ws.cell(row=row, column=7, value=values['oilGallons'] * OIL_PRICE).border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    # Total
    total = (values['labor'] + values['mobilization']) * LABOR_RATE + values['filterCost'] + (values['oilGallons'] * OIL_PRICE)
    ws.cell(row=row, column=8, value=total).border = border
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True, size=10)

    ws.cell(row=row, column=9, value="1x per year").border = border
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=9).font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

# Pricing note
row += 1
ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Pricing Details: Labor at ${LABOR_RATE:.2f}/hr. Oil pricing: ${OIL_BASE:.2f}/gal base cost × {OIL_MARKUP}x markup = ${OIL_PRICE:.2f}/gal (includes overhead & handling)."
cell.fill = note_fill
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 30
row += 3

# ============================================================================
# SERVICE C - COOLANT SERVICE
# ============================================================================
print("Building Service C...")
service_c = settings['serviceC']

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE C - COOLANT SERVICE"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Coolant System Service - Frequency: {service_c.get('frequency', 'As Needed')}"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

# Column headers
headers_c = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Coolant Volume\nNeeded\n(gallons)",
    "Coolant Cost\n@ $"+f"{COOLANT_PRICE:.2f}"+"/gallon\n(with markup)",
    "Hoses/Belts\nReplacement\nCost",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col, header in enumerate(headers_c, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.border = border
    cell.fill = col_header_fill
    cell.font = col_header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[chr(64+col)].width = [14, 18, 12, 12, 12, 12, 12, 14, 14][col-1]
ws.row_dimensions[row].height = 45
row += 1

# Data rows
for kw_range, values in service_c['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).fill = category_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=3).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=5, value=values['coolantGallons']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=5).number_format = '0.0 "gal"'

    ws.cell(row=row, column=6, value=values['coolantGallons'] * COOLANT_PRICE).border = border
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=6).number_format = '$#,##0.00'

    ws.cell(row=row, column=7, value=values['hoseBeltCost']).border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    # Total
    total = (values['labor'] + values['mobilization']) * LABOR_RATE + (values['coolantGallons'] * COOLANT_PRICE) + values['hoseBeltCost']
    ws.cell(row=row, column=8, value=total).border = border
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True, size=10)

    ws.cell(row=row, column=9, value="As needed").border = border
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=9).font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

# Pricing note
row += 1
ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Pricing Details: Labor at ${LABOR_RATE:.2f}/hr. Coolant pricing: ${COOLANT_BASE:.2f}/gal base cost × {COOLANT_MARKUP}x markup = ${COOLANT_PRICE:.2f}/gal (includes overhead & handling)."
cell.fill = note_fill
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 30
row += 3

# ============================================================================
# SERVICE D - LAB ANALYSIS
# ============================================================================
print("Building Service D...")
service_d = settings['serviceD']

ws.merge_cells(f'A{row}:C{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE D - LABORATORY ANALYSIS"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:C{row}')
cell = ws[f'A{row}']
cell.value = "Fluid Analysis Services - Frequency: As Needed"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

# Headers
ws.cell(row=row, column=1, value="Analysis Type").border = border
ws.cell(row=row, column=1).fill = col_header_fill
ws.cell(row=row, column=1).font = col_header_font
ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.column_dimensions['A'].width = 25

ws.cell(row=row, column=2, value="Laboratory Fee\nper Sample").border = border
ws.cell(row=row, column=2).fill = col_header_fill
ws.cell(row=row, column=2).font = col_header_font
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.column_dimensions['B'].width = 18

ws.cell(row=row, column=3, value="Notes").border = border
ws.cell(row=row, column=3).fill = col_header_fill
ws.cell(row=row, column=3).font = col_header_font
ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
ws.column_dimensions['C'].width = 40
ws.row_dimensions[row].height = 30
row += 1

# Oil Analysis
ws.cell(row=row, column=1, value="Oil Analysis").border = border
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=row, column=2, value=service_d['oilAnalysisCost']).border = border
ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=2).number_format = '$#,##0.00'
ws.cell(row=row, column=3, value="Standard oil analysis (wear metals, contamination)").border = border
ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 18
row += 1

# Coolant Analysis
ws.cell(row=row, column=1, value="Coolant Analysis").border = border
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=row, column=2, value=service_d['coolantAnalysisCost']).border = border
ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=2).number_format = '$#,##0.00'
ws.cell(row=row, column=3, value="Coolant condition & additive analysis").border = border
ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 18
row += 1

# Fuel Analysis
ws.cell(row=row, column=1, value="Fuel Analysis").border = border
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=row, column=2, value=service_d['fuelAnalysisCost']).border = border
ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=2).number_format = '$#,##0.00'
ws.cell(row=row, column=3, value="Fuel quality, contamination, microbial growth").border = border
ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 18
row += 1

# Pricing note
row += 1
ws.merge_cells(f'A{row}:C{row}')
cell = ws[f'A{row}']
cell.value = f"Pricing Details: Laboratory fees only. No mobilization required (samples collected during other services). Pricing from calculator settings."
cell.fill = note_fill
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 30
row += 3

# ============================================================================
# SERVICE E - LOAD BANK TESTING
# ============================================================================
print("Building Service E...")
service_e = settings['serviceE']

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE E - LOAD BANK TESTING"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "Full-Load Performance Testing - Frequency: Annual"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

# Column headers
headers_e = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Load Bank\nEquipment\nRental",
    "Transformer\nRental\n(if needed)",
    "Delivery &\nSetup\nCost",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col, header in enumerate(headers_e, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.border = border
    cell.fill = col_header_fill
    cell.font = col_header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[chr(64+col)].width = [14, 18, 12, 12, 12, 12, 12, 14, 14][col-1]
ws.row_dimensions[row].height = 45
row += 1

# Data rows
for kw_range, values in service_e['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).fill = category_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=3).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=5, value=values['loadBankRental']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=5).number_format = '$#,##0.00'

    ws.cell(row=row, column=6, value=values['transformerRental']).border = border
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=6).number_format = '$#,##0.00'

    ws.cell(row=row, column=7, value=values['deliveryCost']).border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    # Total
    total = (values['labor'] + values['mobilization']) * LABOR_RATE + values['loadBankRental'] + values['transformerRental'] + values['deliveryCost']
    ws.cell(row=row, column=8, value=total).border = border
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True, size=10)

    ws.cell(row=row, column=9, value="1x per year").border = border
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=9).font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

# Pricing note
row += 1
ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Pricing Details: Labor at ${LABOR_RATE:.2f}/hr. Equipment rental and delivery costs from calculator settings. Large units may require transformer rental for voltage matching."
cell.fill = note_fill
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 30

# Save
output_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-Formula-Driven-FINAL-LABELED-All-36-Units.xlsx"
wb.save(output_path)

print(f"\n[OK] Complete! All pricing from default-settings.json")
print(f"     Labor Rate: ${LABOR_RATE:.2f}/hr")
print(f"     Oil: ${OIL_BASE:.2f}/gal × {OIL_MARKUP}x = ${OIL_PRICE:.2f}/gal")
print(f"     Coolant: ${COOLANT_BASE:.2f}/gal × {COOLANT_MARKUP}x = ${COOLANT_PRICE:.2f}/gal")
print(f"     File: {output_path}")
