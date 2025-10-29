import openpyxl
import json

wb = openpyxl.load_workbook(r'G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\5-Offerors Pricing.xlsx')
sheet = wb['Pricing']

structure = {
    "dimensions": f"{sheet.dimensions}",
    "max_row": sheet.max_row,
    "max_column": sheet.max_column,
    "sections": {}
}

# Row 11: Year headers
year_headers = {}
for col_idx in range(1, 14):
    val = sheet.cell(11, col_idx).value
    if val:
        year_headers[f"col_{col_idx}"] = val
structure["year_headers_row_11"] = year_headers

# Row 12: Column headers
col_headers = {}
for col_idx in range(1, 14):
    val = sheet.cell(12, col_idx).value
    if val:
        col_headers[f"col_{col_idx}"] = val
structure["column_headers_row_12"] = col_headers

# Section A: Annual PM (rows 13-48)
section_a_generators = []
for row_idx in range(13, 49):
    gen_num = sheet.cell(row_idx, 2).value
    if gen_num:
        section_a_generators.append({
            "row": row_idx,
            "generator": gen_num,
            "year1_2_price_col": 3,
            "year3_price_col": 6,
            "year4_price_col": 9,
            "total_price_col": 12
        })
structure["sections"]["A_annual_pm"] = {
    "start_row": 13,
    "end_row": 48,
    "generators": section_a_generators
}

# Section B: Hourly rates for repairs (rows 51-52)
structure["sections"]["B_repair_rates"] = {
    "electrician_straight_time": {"row": 51, "price_col": 3},
    "electrician_overtime": {"row": 52, "price_col": 3}
}

# Section C: Emergency repair rates (rows 55-56)
structure["sections"]["C_emergency_rates"] = {
    "electrician_straight_time": {"row": 55, "price_col": 3},
    "electrician_overtime": {"row": 56, "price_col": 3}
}

# Section D: Other services (row 59)
structure["sections"]["D_other_services"] = {
    "fuel_sample": {"row": 59, "price_col": 3}
}

# Section E: Parts pricing (row 61)
structure["sections"]["E_parts"] = {
    "parts_description": {"row": 61, "description_col": 3}
}

print(json.dumps(structure, indent=2))
