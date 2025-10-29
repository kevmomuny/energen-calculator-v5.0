"""
Rebuild Service Lookup Tables with proper labeling for LBNL RFP
Each row will have clear, descriptive labels
"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load the workbook
excel_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-Formula-Driven-FINAL-All-36-Units.xlsx"
wb = load_workbook(excel_path)

# Load service data from Calculator
with open(r"C:\ECalc\active\energen-calculator-v5.0\frontend\config\default-settings.json", 'r') as f:
    calc_settings = json.load(f)

# Delete and recreate Service Lookup Tables sheet
if "Service Lookup Tables" in wb.sheetnames:
    del wb["Service Lookup Tables"]

ws = wb.create_sheet("Service Lookup Tables", 1)

# Styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, size=14, color="FFFFFF")
subtitle_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subtitle_font = Font(bold=True, size=11, color="FFFFFF")
table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
table_header_font = Font(bold=True, size=10)
size_label_font = Font(bold=True, size=10, color="000000")
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
note_font = Font(bold=True, size=10, color="C65911")

border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Tier descriptions
tier_names = {
    "2-14": "Very Small Generators",
    "15-30": "Small Generators",
    "35-150": "Medium Generators",
    "151-250": "Large Generators",
    "251-400": "Very Large Generators",
    "401-500": "Extra Large Generators",
    "501-670": "Industrial Generators",
    "671-1050": "Large Industrial Generators",
    "1051-1500": "Very Large Industrial",
    "1501+": "Extra Large Industrial"
}

row = 1

# ============================================================================
# SERVICE A
# ============================================================================
print("Creating Service A table...")
service_a = calc_settings['serviceA']

# Header
ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE A - COMPREHENSIVE INSPECTION"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
row += 1

# Subtitle
ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = f"{service_a['name']} | Frequency: {service_a['frequency']}× per year (Quarterly)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
row += 1

# Column headers
headers_a = ["kW Range", "Generator Size Category", "Labor\n(hours)", "Mobilization\n(hours)", "Parts Cost\n(inspection supplies)", "Total Cost\nper Visit", "Notes"]
for col, header in enumerate(headers_a, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = table_header_font
    cell.fill = table_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[row].height = 40
row += 1

# Data rows
for kw_range, values in service_a['data'].items():
    # kW Range
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    # Size Category
    ws.cell(row=row, column=2, value=tier_names.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=2).font = size_label_font

    # Labor hours
    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3).number_format = '0.0'

    # Mobilization hours
    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).number_format = '0.0'

    # Parts cost
    ws.cell(row=row, column=5, value=values['parts']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5).number_format = '$#,##0.00'

    # Total formula
    formula = f"=(C{row}+D{row})*'RFP Parameters'!$B$11+E{row}"
    cell = ws.cell(row=row, column=6, value=formula)
    cell.border = border
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = '$#,##0.00'
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cell.font = Font(bold=True, size=10)

    # Notes
    ws.cell(row=row, column=7, value="Quarterly service").border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=7).font = Font(size=9, italic=True)

    row += 1

# Formula explanation
row += 1
ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = "Formula: Total = (Labor + Mobilization hours) × $311.24/hr Prevailing Wage + Parts"
cell.font = Font(italic=True, size=9, color="666666")
row += 2

# Set column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 18

print("  [OK] Service A complete")

# ============================================================================
# SERVICE B
# ============================================================================
print("Creating Service B table...")
service_b = calc_settings['serviceB']

ws.merge_cells(f'A{row}:H{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE B - OIL & FILTER SERVICE"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
row += 1

ws.merge_cells(f'A{row}:H{row}')
cell = ws[f'A{row}']
cell.value = f"{service_b['name']} | Frequency: {service_b['frequency']}× per year (Annual)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
row += 1

headers_b = ["kW Range", "Generator Size Category", "Labor\n(hours)", "Mobilization\n(hours)", "Oil Filter\nCost", "Oil Volume\n(gallons)", "Oil Cost\n@ $24/gal", "Total Cost\nper Visit"]
for col, header in enumerate(headers_b, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = table_header_font
    cell.fill = table_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[row].height = 40
row += 1

oil_price = calc_settings['oilPrice'] * calc_settings['oilMarkup']
for kw_range, values in service_b['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=tier_names.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=2).font = size_label_font

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3).number_format = '0.0'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).number_format = '0.0'

    ws.cell(row=row, column=5, value=values['filterCost']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5).number_format = '$#,##0.00'

    ws.cell(row=row, column=6, value=values['oilGallons']).border = border
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=6).number_format = '0.0'

    ws.cell(row=row, column=7, value=values['oilGallons'] * oil_price).border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    formula = f"=(C{row}+D{row})*'RFP Parameters'!$B$11+E{row}+G{row}"
    cell = ws.cell(row=row, column=8, value=formula)
    cell.border = border
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = '$#,##0.00'
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cell.font = Font(bold=True, size=10)

    row += 1

row += 1
ws.merge_cells(f'A{row}:H{row}')
cell = ws[f'A{row}']
cell.value = f"Oil Pricing: ${calc_settings['oilPrice']:.2f}/gal base × {calc_settings['oilMarkup']} markup = ${oil_price:.2f}/gal"
cell.font = note_font
cell.fill = note_fill
cell.alignment = Alignment(horizontal="center")
row += 2

ws.column_dimensions['H'].width = 16

print("  ✓ Service B complete")

# Continue with Services C, D, E...
# (Similar pattern for remaining services)

print("\nSaving workbook...")
wb.save(excel_path.replace("FINAL", "FINAL-LABELED"))
wb.close()
print(f"✓ Saved: LBNL-Formula-Driven-FINAL-LABELED-All-36-Units.xlsx")
