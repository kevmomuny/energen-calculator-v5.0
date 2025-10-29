#!/usr/bin/env python3
"""
Phase 5: Fill Excel Bid Template with Calculator-Verified Pricing
"""
import openpyxl
import json
from datetime import datetime
from openpyxl.styles import numbers

RFP_DIR = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP"
SOURCE_DIR = f"{RFP_DIR}\\Request for Proposal No. ANR-6-2025 Genertor Services"
OUTPUT_DIR = f"{RFP_DIR}\\RFP-Evaluator"

print("\n" + "="*80)
print("PHASE 5: EXCEL TEMPLATE FILL WITH VERIFIED PRICING")
print("="*80)

# Load pricing results
print("\nLoading verified pricing data...")
with open(f"{OUTPUT_DIR}/rfp-pricing-results.json", 'r') as f:
    pricing_data = json.load(f)

print(f"  OK Loaded {len(pricing_data['generators'])} generators")
print(f"  OK Annual total: ${pricing_data['summary']['annualTotal']:,.2f}")

# Create lookup dictionary by unit number
pricing_lookup = {}
for gen in pricing_data['generators']:
    unit_num = gen['unitNumber'].strip().upper()
    pricing_lookup[unit_num] = gen['annualPrice']

print(f"  OK Created pricing lookup for {len(pricing_lookup)} generators")

# Load Excel template
template_path = f"{SOURCE_DIR}/5-Offerors Pricing.xlsx"
print(f"\nLoading Excel template...")
print(f"  Path: {template_path}")

wb = openpyxl.load_workbook(template_path)
ws = wb.active

print(f"  OK Sheet: {ws.title}")
print(f"  OK Max row: {ws.max_row}")

# Fill in company info
ws['A5'] = "Supplier: Energen Systems Inc."
ws['A6'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"

# Find pricing rows (row 13 onwards based on structure analysis)
filled_count = 0
not_found_count = 0
not_found_list = []

print("\nFilling pricing data...")
for row in range(13, 52):  # Rows 13-51 (39 generators)
    unit_number = ws.cell(row, 2).value  # Column B has unit numbers

    if unit_number:
        unit_num_clean = str(unit_number).strip().upper()

        if unit_num_clean in pricing_lookup:
            price = pricing_lookup[unit_num_clean]
            ws.cell(row, 3).value = price  # Column C for price
            ws.cell(row, 3).number_format = '$#,##0.00'
            filled_count += 1
            print(f"  [{filled_count:2d}] {unit_number:20s} -> ${price:>10,.2f}")
        else:
            not_found_count += 1
            not_found_list.append(unit_number)
            print(f"  [XX] {unit_number:20s} -> NOT FOUND in pricing data")

print(f"\n{'='*80}")
print("FILL SUMMARY")
print(f"{'='*80}")
print(f"\nFilled: {filled_count}")
print(f"Not found: {not_found_count}")

if not_found_list:
    print(f"\nGenerators not found:")
    for unit in not_found_list:
        print(f"  - {unit}")

# Save filled template
output_path = f"{OUTPUT_DIR}/LBNL-5-Offerors-Pricing-FILLED.xlsx"
wb.save(output_path)
print(f"\nOK Saved: LBNL-5-Offerors-Pricing-FILLED.xlsx")

# Also create a CSV summary for quick reference
print("\nCreating CSV summary...")
import csv

csv_path = f"{OUTPUT_DIR}/pricing-summary.csv"
with open(csv_path, 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['Unit Number', 'kW Rating', 'Annual Price', 'Labor', 'Parts', 'Mobilization', 'Tax'])

    for gen in pricing_data['generators']:
        if gen.get('success', False):
            writer.writerow([
                gen['unitNumber'],
                gen['kwRating'],
                f"${gen['annualPrice']:,.2f}",
                f"${gen['breakdown']['labor']:,.2f}",
                f"${gen['breakdown']['parts']:,.2f}",
                f"${gen['breakdown']['mobilization']:,.2f}",
                f"${gen['breakdown']['tax']:,.2f}"
            ])

print(f"  OK Saved: pricing-summary.csv")

print("\n" + "="*80)
print("PHASE 5 COMPLETE - EXCEL TEMPLATE FILLED")
print("="*80)
print("\nDeliverables:")
print(f"  1. LBNL-5-Offerors-Pricing-FILLED.xlsx")
print(f"  2. pricing-summary.csv")
print(f"\nNext: Phase 6 - Package complete bid deliverables")
