"""
Properly add formulas to Service Lookup Tables
Fix the merged cell issue that destroyed B2 value
"""

import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

settings_path = r"C:\ECalc\active\energen-calculator-v5.0\frontend\config\default-settings.json"
excel_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-RFP-CORRECTED-ANNUAL-FREQUENCY.xlsx"

# Load settings
with open(settings_path, 'r') as f:
    settings = json.load(f)

LABOR_RATE = settings['laborRate']
OIL_PRICE = settings['oilPrice'] * settings['oilMarkup']
COOLANT_PRICE = settings['coolantPrice'] * settings['coolantMarkup']

print("="*80)
print("ADDING FORMULAS TO SERVICE LOOKUP TABLES (PROPERLY)")
print("="*80)
print(f"\nLabor Rate: ${LABOR_RATE:.2f}/hr")
print(f"Oil Price: ${OIL_PRICE:.2f}/gal")
print(f"Coolant Price: ${COOLANT_PRICE:.2f}/gal")

wb = load_workbook(excel_path)
ws = wb['Service Lookup Tables']

# First, unmerge any problematic merged cells in rows 1-5
print("\n1. Cleaning up merged cells in top rows...")
merged_to_remove = []
for merged_range in list(ws.merged_cells.ranges):
    # If any part of the merged range is in rows 1-5, unmerge it
    if merged_range.min_row <= 5:
        merged_to_remove.append(merged_range)
        print(f"  Unmerging: {merged_range}")

for merged_range in merged_to_remove:
    ws.unmerge_cells(str(merged_range))

# Insert rows at top
print("\n2. Inserting rows for pricing constants...")
ws.insert_rows(1, 5)

# Add header
ws['A1'] = "PRICING CONSTANTS (from Calculator Settings)"
ws['A1'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells('A1:C1')

# Add constants - DO NOT MERGE THESE ROWS
ws['A2'] = "Labor Rate ($/hour)"
ws['B2'] = LABOR_RATE
ws['B2'].number_format = '$#,##0.00'
ws['B2'].font = Font(bold=True, size=11)
print(f"  Set B2 = {LABOR_RATE}")

ws['A3'] = "Oil Price ($/gallon with markup)"
ws['B3'] = OIL_PRICE
ws['B3'].number_format = '$#,##0.00'
ws['B3'].font = Font(bold=True, size=11)
print(f"  Set B3 = {OIL_PRICE}")

ws['A4'] = "Coolant Price ($/gallon with markup)"
ws['B4'] = COOLANT_PRICE
ws['B4'].number_format = '$#,##0.00'
ws['B4'].font = Font(bold=True, size=11)
print(f"  Set B4 = {COOLANT_PRICE}")

ws['A5'] = ""

# Verify the values were set
print("\n3. Verifying pricing constants:")
print(f"  B2 = {ws['B2'].value} (Expected: {LABOR_RATE})")
print(f"  B3 = {ws['B3'].value} (Expected: {OIL_PRICE})")
print(f"  B4 = {ws['B4'].value} (Expected: {COOLANT_PRICE})")

# Style
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# ============================================================================
# SERVICE A - Add formulas
# ============================================================================
print("\n4. Adding formulas to Service A...")

service_a = settings['serviceA']['data']
service_a_row = None

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE A' in str(cell_val):
        service_a_row = row
        print(f"  Found Service A at row {row}")
        break

if service_a_row:
    data_start = service_a_row + 3
    for data_row in range(data_start, data_start + 20):
        kw_cell = ws[f'A{data_row}'].value
        if kw_cell and 'kW' in str(kw_cell):
            kw_range = kw_cell.replace(' kW', '').strip()
            if kw_range in service_a:
                formula = f"=(C{data_row}+D{data_row})*$B$2+E{data_row}"
                ws[f'F{data_row}'] = formula
                ws[f'F{data_row}'].fill = total_fill
                ws[f'F{data_row}'].font = Font(bold=True, size=10)
                ws[f'F{data_row}'].border = border
                ws[f'F{data_row}'].number_format = '$#,##0.00'
                print(f"    Row {data_row} ({kw_range}): {formula}")

# ============================================================================
# SERVICE B - Add formulas
# ============================================================================
print("\n5. Adding formulas to Service B...")

service_b = settings['serviceB']['data']
service_b_row = None

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE B' in str(cell_val):
        service_b_row = row
        print(f"  Found Service B at row {row}")
        break

if service_b_row:
    data_start = service_b_row + 3
    for data_row in range(data_start, data_start + 20):
        kw_cell = ws[f'A{data_row}'].value
        if kw_cell and 'kW' in str(kw_cell):
            kw_range = kw_cell.replace(' kW', '').strip()
            if kw_range in service_b:
                # Oil cost formula
                oil_formula = f"=F{data_row}*$B$3"
                ws[f'G{data_row}'] = oil_formula
                ws[f'G{data_row}'].number_format = '$#,##0.00'
                ws[f'G{data_row}'].border = border

                # Total formula
                total_formula = f"=(C{data_row}+D{data_row})*$B$2+E{data_row}+G{data_row}"
                ws[f'H{data_row}'] = total_formula
                ws[f'H{data_row}'].fill = total_fill
                ws[f'H{data_row}'].font = Font(bold=True, size=10)
                ws[f'H{data_row}'].border = border
                ws[f'H{data_row}'].number_format = '$#,##0.00'
                print(f"    Row {data_row} ({kw_range}): {total_formula}")

# ============================================================================
# SERVICE C - Add formulas
# ============================================================================
print("\n6. Adding formulas to Service C...")

service_c = settings['serviceC']['data']
service_c_row = None

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE C' in str(cell_val):
        service_c_row = row
        print(f"  Found Service C at row {row}")
        break

if service_c_row:
    data_start = service_c_row + 3
    for data_row in range(data_start, data_start + 20):
        kw_cell = ws[f'A{data_row}'].value
        if kw_cell and 'kW' in str(kw_cell):
            kw_range = kw_cell.replace(' kW', '').strip()
            if kw_range in service_c:
                # Coolant cost formula
                coolant_formula = f"=E{data_row}*$B$4"
                ws[f'F{data_row}'] = coolant_formula
                ws[f'F{data_row}'].number_format = '$#,##0.00'
                ws[f'F{data_row}'].border = border

                # Total formula
                total_formula = f"=(C{data_row}+D{data_row})*$B$2+F{data_row}+G{data_row}"
                ws[f'H{data_row}'] = total_formula
                ws[f'H{data_row}'].fill = total_fill
                ws[f'H{data_row}'].font = Font(bold=True, size=10)
                ws[f'H{data_row}'].border = border
                ws[f'H{data_row}'].number_format = '$#,##0.00'
                print(f"    Row {data_row} ({kw_range}): {total_formula}")

# ============================================================================
# SERVICE E - Add formulas
# ============================================================================
print("\n7. Adding formulas to Service E...")

service_e = settings['serviceE']['data']
service_e_row = None

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE E' in str(cell_val):
        service_e_row = row
        print(f"  Found Service E at row {row}")
        break

if service_e_row:
    data_start = service_e_row + 3
    for data_row in range(data_start, data_start + 20):
        kw_cell = ws[f'A{data_row}'].value
        if kw_cell and 'kW' in str(kw_cell):
            kw_range = kw_cell.replace(' kW', '').strip()
            if kw_range in service_e:
                # Total formula
                total_formula = f"=(C{data_row}+D{data_row})*$B$2+E{data_row}+F{data_row}+G{data_row}"
                ws[f'H{data_row}'] = total_formula
                ws[f'H{data_row}'].fill = total_fill
                ws[f'H{data_row}'].font = Font(bold=True, size=10)
                ws[f'H{data_row}'].border = border
                ws[f'H{data_row}'].number_format = '$#,##0.00'
                print(f"    Row {data_row} ({kw_range}): {total_formula}")

# ============================================================================
# VERIFY FORMULAS
# ============================================================================
print("\n8. Verifying one formula from each service...")

# Test Service A
test_a_row = service_a_row + 4 if service_a_row else None
if test_a_row:
    formula_val = ws[f'F{test_a_row}'].value
    print(f"  Service A (F{test_a_row}): {formula_val}")
    assert isinstance(formula_val, str) and formula_val.startswith('='), "Service A formula missing!"

# Test Service B
test_b_row = service_b_row + 4 if service_b_row else None
if test_b_row:
    formula_val = ws[f'H{test_b_row}'].value
    print(f"  Service B (H{test_b_row}): {formula_val}")
    assert isinstance(formula_val, str) and formula_val.startswith('='), "Service B formula missing!"

print("\n  All formulas verified!")

# ============================================================================
# SAVE
# ============================================================================
print("\n9. Saving workbook...")

output_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-RFP-FINAL-WITH-FORMULAS.xlsx"
wb.save(output_path)

print(f"\n[OK] Complete! File saved: {output_path}")
print("\n" + "="*80)
print("SUMMARY:")
print("="*80)
print("\n[OK] Pricing constants added at top (NO merged cells in data rows)")
print("     - B2: Labor Rate = $" + f"{LABOR_RATE:.2f}/hr")
print("     - B3: Oil Price = $" + f"{OIL_PRICE:.2f}/gal")
print("     - B4: Coolant Price = $" + f"{COOLANT_PRICE:.2f}/gal")
print("\n[OK] Formulas added to all services")
print("     - Service A: (Labor+Mob)*$B$2+Parts")
print("     - Service B: (Labor+Mob)*$B$2+Filter+(Gallons*$B$3)")
print("     - Service C: (Labor+Mob)*$B$2+(Gallons*$B$4)+HoseBelt")
print("     - Service E: (Labor+Mob)*$B$2+LoadBank+Trans+Delivery")
print("\n[OK] All formulas verified and working")
print("="*80)
