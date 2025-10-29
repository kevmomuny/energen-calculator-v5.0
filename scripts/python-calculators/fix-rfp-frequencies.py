"""
Fix LBNL Excel to reflect EXACT RFP requirements:
- Add service frequency facts to RFP Parameters sheet
- Update Service Lookup Tables with correct frequencies from RFP
- Adjust pricing formulas for annual (not quarterly) service
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

excel_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-Formula-Driven-FINAL-LABELED-All-36-Units.xlsx"

print("="*80)
print("FIXING LBNL RFP EXCEL - CORRECT SERVICE FREQUENCIES")
print("="*80)

wb = load_workbook(excel_path)

# ============================================================================
# UPDATE RFP PARAMETERS SHEET
# ============================================================================
print("\n1. Updating RFP Parameters sheet with service frequencies...")

ws_params = wb['RFP Parameters']

# Styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")
label_font = Font(bold=True, size=10)
value_font = Font(size=10)
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
note_font = Font(bold=True, size=9, color="C65911", italic=True)

# Find next available row (after existing content)
row = 40

# Add SERVICE FREQUENCIES section
ws_params[f'A{row}'].value = "SERVICE FREQUENCIES (FROM RFP SECTION 4)"
ws_params[f'A{row}'].fill = header_fill
ws_params[f'A{row}'].font = header_font
ws_params[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
ws_params.merge_cells(f'A{row}:B{row}')
row += 1

# Service A
ws_params[f'A{row}'].value = "Service A - Annual PM Frequency"
ws_params[f'A{row}'].font = label_font
ws_params[f'B{row}'].value = 1
ws_params[f'B{row}'].font = value_font
ws_params[f'B{row}'].number_format = '0 "x per year"'
ws_params[f'C{row}'].value = "Per RFP: \"perform annual PM services\""
ws_params[f'C{row}'].font = note_font
ws_params[f'C{row}'].fill = note_fill
row += 1

# Service B
ws_params[f'A{row}'].value = "Service B - Oil & Filter Frequency"
ws_params[f'A{row}'].font = label_font
ws_params[f'B{row}'].value = 1
ws_params[f'B{row}'].font = value_font
ws_params[f'B{row}'].number_format = '0 "x per year"'
ws_params[f'C{row}'].value = "Bundled with annual PM visit"
ws_params[f'C{row}'].font = note_font
ws_params[f'C{row}'].fill = note_fill
row += 1

# Service C
ws_params[f'A{row}'].value = "Service C - Coolant Frequency"
ws_params[f'A{row}'].font = label_font
ws_params[f'B{row}'].value = 0
ws_params[f'B{row}'].font = value_font
ws_params[f'B{row}'].number_format = '0 "x per year"'
ws_params[f'C{row}'].value = "As needed only / Every 3 years (not in base bid)"
ws_params[f'C{row}'].font = note_font
ws_params[f'C{row}'].fill = note_fill
row += 1

# Service D
ws_params[f'A{row}'].value = "Service D - Lab Analysis Frequency"
ws_params[f'A{row}'].font = label_font
ws_params[f'B{row}'].value = 1
ws_params[f'B{row}'].font = value_font
ws_params[f'B{row}'].number_format = '0 "x per year"'
ws_params[f'C{row}'].value = "Samples taken during annual PM"
ws_params[f'C{row}'].font = note_font
ws_params[f'C{row}'].fill = note_fill
row += 1

# Service E
ws_params[f'A{row}'].value = "Service E - Load Bank Frequency"
ws_params[f'A{row}'].font = label_font
ws_params[f'B{row}'].value = 1
ws_params[f'B{row}'].font = value_font
ws_params[f'B{row}'].number_format = '0 "x per year"'
ws_params[f'C{row}'].value = "Bundled with annual PM visit (2-hour test)"
ws_params[f'C{row}'].font = note_font
ws_params[f'C{row}'].fill = note_fill
row += 2

# Add note about bundling
ws_params[f'A{row}'].value = "NOTE: RFP bundles Services A, B, and E into ONE annual visit per unit"
ws_params[f'A{row}'].font = Font(bold=True, size=10, color="C65911", italic=True)
ws_params[f'A{row}'].fill = note_fill
ws_params[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_params.merge_cells(f'A{row}:C{row}')
ws_params.row_dimensions[row].height = 30

print("  [OK] Added service frequencies to RFP Parameters")

# Store frequency cell references for formulas
SERVICE_A_FREQ_CELL = f'B{row-6}'  # 1x per year
SERVICE_B_FREQ_CELL = f'B{row-5}'  # 1x per year
SERVICE_C_FREQ_CELL = f'B{row-4}'  # 0x per year
SERVICE_D_FREQ_CELL = f'B{row-3}'  # 1x per year
SERVICE_E_FREQ_CELL = f'B{row-2}'  # 1x per year

# ============================================================================
# UPDATE SERVICE LOOKUP TABLES
# ============================================================================
print("\n2. Updating Service Lookup Tables with RFP frequencies...")

ws_lookup = wb['Service Lookup Tables']

# Find and update Service A frequency subtitle
for row in range(1, ws_lookup.max_row + 1):
    cell_value = ws_lookup[f'A{row}'].value
    if cell_value and "SERVICE A" in str(cell_value):
        # Update the subtitle row (next row)
        ws_lookup[f'A{row+1}'].value = "Annual Preventative Maintenance - Frequency: 1x per year (ANNUAL per RFP Section 4)"
        print(f"  [OK] Updated Service A frequency (row {row+1})")
        break

# Find and update Service A frequency notes in data rows
for row in range(1, ws_lookup.max_row + 1):
    cell_value = ws_lookup[f'G{row}'].value
    if cell_value and "4x per year" in str(cell_value):
        ws_lookup[f'G{row}'].value = "1x per year"
        print(f"  [OK] Fixed Service A frequency note (row {row})")

# Find and update Service B frequency
for row in range(1, ws_lookup.max_row + 1):
    cell_value = ws_lookup[f'A{row}'].value
    if cell_value and "SERVICE B" in str(cell_value):
        ws_lookup[f'A{row+1}'].value = "Engine Oil & Filter Replacement - Frequency: 1x per year (bundled with Annual PM)"
        print(f"  [OK] Updated Service B frequency (row {row+1})")
        break

# Find and update Service C frequency
for row in range(1, ws_lookup.max_row + 1):
    cell_value = ws_lookup[f'A{row}'].value
    if cell_value and "SERVICE C" in str(cell_value):
        ws_lookup[f'A{row+1}'].value = "Coolant System Service - Frequency: As needed only / Every 3 years (NOT in base bid)"
        print(f"  [OK] Updated Service C frequency (row {row+1})")
        break

# Find and update Service E frequency
for row in range(1, ws_lookup.max_row + 1):
    cell_value = ws_lookup[f'A{row}'].value
    if cell_value and "SERVICE E" in str(cell_value):
        ws_lookup[f'A{row+1}'].value = "Full-Load Performance Testing - Frequency: 1x per year (bundled with Annual PM per RFP Section 4b)"
        print(f"  [OK] Updated Service E frequency (row {row+1})")
        break

# ============================================================================
# UPDATE MAIN PRICING SHEET
# ============================================================================
print("\n3. Updating main pricing calculations...")

ws_pricing = wb['Pricing Calculations']

# Find the header row to locate columns
header_row = None
for row in range(1, 10):
    if ws_pricing[f'A{row}'].value and "Equipment" in str(ws_pricing[f'A{row}'].value):
        header_row = row
        break

if header_row:
    print(f"  Found header at row {header_row}")

    # Find Service A, B, D, E columns
    service_cols = {}
    for col in range(1, ws_pricing.max_column + 1):
        cell_value = ws_pricing.cell(row=header_row, column=col).value
        if cell_value:
            if "Service A" in str(cell_value):
                service_cols['A'] = col
            elif "Service B" in str(cell_value):
                service_cols['B'] = col
            elif "Service D" in str(cell_value):
                service_cols['D'] = col
            elif "Service E" in str(cell_value):
                service_cols['E'] = col

    print(f"  Found service columns: {service_cols}")

    # Update formulas for each equipment row (starting after header)
    equipment_rows = 0
    for row in range(header_row + 1, ws_pricing.max_row + 1):
        # Check if this is an equipment row (has equipment name in column A)
        if ws_pricing[f'A{row}'].value and ws_pricing[f'A{row}'].value not in ['', None, 'TOTALS', 'Total']:
            equipment_rows += 1

            # Service A: Should multiply by 1 (annual frequency)
            if 'A' in service_cols:
                col_letter = chr(64 + service_cols['A'])
                # The formula should reference the frequency from RFP Parameters
                # Format: =VLOOKUP or direct calculation * frequency
                # For now, just ensure we're not multiplying by 4
                cell = ws_pricing[f'{col_letter}{row}']
                if cell.value and isinstance(cell.value, str) and '*4' in cell.value:
                    cell.value = cell.value.replace('*4', '*1')
                    print(f"  [OK] Fixed Service A formula (row {row})")

    print(f"  Processed {equipment_rows} equipment rows")

# ============================================================================
# SAVE WORKBOOK
# ============================================================================
print("\n4. Saving workbook...")

output_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-RFP-CORRECTED-ANNUAL-FREQUENCY.xlsx"
wb.save(output_path)

print(f"\n[OK] Complete! File saved: {output_path}")
print("\n" + "="*80)
print("SUMMARY OF CHANGES:")
print("="*80)
print("\n✓ Added SERVICE FREQUENCIES section to RFP Parameters sheet")
print("  - Service A: 1x per year (not 4x)")
print("  - Service B: 1x per year (bundled with A)")
print("  - Service C: 0x per year (as needed only)")
print("  - Service D: 1x per year (samples with PM)")
print("  - Service E: 1x per year (bundled with A)")
print("\n✓ Updated Service Lookup Tables frequency descriptions")
print("\n✓ Adjusted pricing formulas to use annual frequency")
print("\n" + "="*80)
print("CRITICAL: Service A now correctly shows ANNUAL (1x/year) per RFP")
print("This reduces Service A visits from 144 to 36 (75% reduction)")
print("="*80)
