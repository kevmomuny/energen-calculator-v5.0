"""
Add actual Excel formulas to Service Lookup Tables
Replace hardcoded values with formulas that reference:
- Labor rate from default-settings.json
- Component costs from each service tier
"""

import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

excel_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-RFP-CORRECTED-ANNUAL-FREQUENCY.xlsx"
settings_path = r"C:\ECalc\active\energen-calculator-v5.0\frontend\config\default-settings.json"

# Load settings
with open(settings_path, 'r') as f:
    settings = json.load(f)

LABOR_RATE = settings['laborRate']
OIL_PRICE = settings['oilPrice'] * settings['oilMarkup']
COOLANT_PRICE = settings['coolantPrice'] * settings['coolantMarkup']

print("="*80)
print("ADDING FORMULAS TO SERVICE LOOKUP TABLES")
print("="*80)
print(f"\nLabor Rate: ${LABOR_RATE:.2f}/hr")
print(f"Oil Price: ${OIL_PRICE:.2f}/gal")
print(f"Coolant Price: ${COOLANT_PRICE:.2f}/gal")

wb = load_workbook(excel_path)
ws = wb['Service Lookup Tables']

# We need to add a reference section at the top for pricing constants
# Insert rows at the top
print("\n1. Adding pricing reference section...")
ws.insert_rows(1, 5)

# Add header
ws['A1'] = "PRICING CONSTANTS (from Calculator Settings)"
ws['A1'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells('A1:C1')

# Add constants
ws['A2'] = "Labor Rate ($/hour)"
ws['B2'] = LABOR_RATE
ws['B2'].number_format = '$#,##0.00'

ws['A3'] = "Oil Price ($/gallon with markup)"
ws['B3'] = OIL_PRICE
ws['B3'].number_format = '$#,##0.00'

ws['A4'] = "Coolant Price ($/gallon with markup)"
ws['B4'] = COOLANT_PRICE
ws['B4'].number_format = '$#,##0.00'

ws['A5'] = ""  # Blank row

print("  [OK] Added pricing constants to rows 1-4")

# Now all row numbers are shifted down by 5
# Service A should start around row 6

# ============================================================================
# SERVICE A - Add formulas
# ============================================================================
print("\n2. Adding formulas to Service A...")

service_a = settings['serviceA']['data']

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE A' in str(cell_val):
        print(f"  Found Service A at row {row}")

        # Find data rows (look for kW values)
        data_start = row + 3  # After header and subtitle

        for data_row in range(data_start, data_start + 20):
            kw_cell = ws[f'A{data_row}'].value
            if kw_cell and 'kW' in str(kw_cell):
                # Extract kW range (e.g., "2-14 kW" -> "2-14")
                kw_range = kw_cell.replace(' kW', '').strip()

                if kw_range in service_a:
                    # Get values
                    labor = service_a[kw_range]['labor']
                    mob = service_a[kw_range]['mobilization']
                    parts = service_a[kw_range]['parts']

                    # Add formula: =(Labor + Mobilization) * LaborRate + Parts
                    # Column C = Labor, D = Mobilization, E = Parts, F = Total
                    formula = f"=(C{data_row}+D{data_row})*$B$2+E{data_row}"
                    ws[f'F{data_row}'] = formula
                    print(f"    Row {data_row} ({kw_range}): {formula}")
        break

# ============================================================================
# SERVICE B - Add formulas
# ============================================================================
print("\n3. Adding formulas to Service B...")

service_b = settings['serviceB']['data']

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE B' in str(cell_val):
        print(f"  Found Service B at row {row}")

        data_start = row + 3

        for data_row in range(data_start, data_start + 20):
            kw_cell = ws[f'A{data_row}'].value
            if kw_cell and 'kW' in str(kw_cell):
                kw_range = kw_cell.replace(' kW', '').strip()

                if kw_range in service_b:
                    # Get values
                    labor = service_b[kw_range]['labor']
                    mob = service_b[kw_range]['mobilization']
                    filter_cost = service_b[kw_range]['filterCost']
                    oil_gallons = service_b[kw_range]['oilGallons']

                    # Add formula: =(Labor + Mob) * LaborRate + FilterCost + (OilGallons * OilPrice)
                    # C=Labor, D=Mob, E=Filter, F=Gallons, G=OilCost, H=Total
                    formula = f"=(C{data_row}+D{data_row})*$B$2+E{data_row}+G{data_row}"
                    ws[f'H{data_row}'] = formula

                    # Also add formula for oil cost (Column G)
                    oil_formula = f"=F{data_row}*$B$3"
                    ws[f'G{data_row}'] = oil_formula

                    print(f"    Row {data_row} ({kw_range}): {formula}")
        break

# ============================================================================
# SERVICE C - Add formulas
# ============================================================================
print("\n4. Adding formulas to Service C...")

service_c = settings['serviceC']['data']

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE C' in str(cell_val):
        print(f"  Found Service C at row {row}")

        data_start = row + 3

        for data_row in range(data_start, data_start + 20):
            kw_cell = ws[f'A{data_row}'].value
            if kw_cell and 'kW' in str(kw_cell):
                kw_range = kw_cell.replace(' kW', '').strip()

                if kw_range in service_c:
                    # Add formula: =(Labor + Mob) * LaborRate + (Gallons * CoolantPrice) + HoseBelt
                    # C=Labor, D=Mob, E=Gallons, F=CoolantCost, G=HoseBelt, H=Total
                    formula = f"=(C{data_row}+D{data_row})*$B$2+F{data_row}+G{data_row}"
                    ws[f'H{data_row}'] = formula

                    # Also add formula for coolant cost (Column F)
                    coolant_formula = f"=E{data_row}*$B$4"
                    ws[f'F{data_row}'] = coolant_formula

                    print(f"    Row {data_row} ({kw_range}): {formula}")
        break

# ============================================================================
# SERVICE E - Add formulas
# ============================================================================
print("\n5. Adding formulas to Service E...")

service_e = settings['serviceE']['data']

for row in range(6, ws.max_row + 1):
    cell_val = ws[f'A{row}'].value
    if cell_val and 'SERVICE E' in str(cell_val):
        print(f"  Found Service E at row {row}")

        data_start = row + 3

        for data_row in range(data_start, data_start + 20):
            kw_cell = ws[f'A{data_row}'].value
            if kw_cell and 'kW' in str(kw_cell):
                kw_range = kw_cell.replace(' kW', '').strip()

                if kw_range in service_e:
                    # Add formula: =(Labor + Mob) * LaborRate + LoadBank + Transformer + Delivery
                    # C=Labor, D=Mob, E=LoadBank, F=Transformer, G=Delivery, H=Total
                    formula = f"=(C{data_row}+D{data_row})*$B$2+E{data_row}+F{data_row}+G{data_row}"
                    ws[f'H{data_row}'] = formula

                    print(f"    Row {data_row} ({kw_range}): {formula}")
        break

# ============================================================================
# SAVE
# ============================================================================
print("\n6. Saving workbook...")

output_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-RFP-WITH-FORMULAS.xlsx"
wb.save(output_path)

print(f"\n[OK] Complete! File saved: {output_path}")
print("\n" + "="*80)
print("SUMMARY:")
print("="*80)
print("\n[OK] Added pricing constants reference at top (rows 1-4)")
print("     - Labor Rate: $B$2")
print("     - Oil Price: $B$3")
print("     - Coolant Price: $B$4")
print("\n[OK] Added formulas to all service tables:")
print("     - Service A: Total = (Labor+Mob)*LaborRate + Parts")
print("     - Service B: Total = (Labor+Mob)*LaborRate + Filter + OilCost")
print("     - Service C: Total = (Labor+Mob)*LaborRate + CoolantCost + HoseBelt")
print("     - Service E: Total = (Labor+Mob)*LaborRate + LoadBank + Trans + Delivery")
print("\n[OK] All totals now dynamically calculate from pricing constants")
print("="*80)
