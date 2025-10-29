"""
Create Service Lookup Tables with COMPLETE labeling
- Descriptive column headers
- Row labels with size categories
- Explanatory notes
"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Paths
excel_path = r"G:\Shared drives\Energen Ops\2-Sales & Marketing\1-Sales\Bids\LNBL-Whole-Facility-RFP\RFP-Evaluator\6th attempt\LBNL-Formula-Driven-FINAL-All-36-Units.xlsx"
settings_path = r"C:\ECalc\active\energen-calculator-v5.0\frontend\config\default-settings.json"

# Load data
with open(settings_path, 'r') as f:
    calc_settings = json.load(f)

wb = load_workbook(excel_path)

# Delete old sheet
if "Service Lookup Tables" in wb.sheetnames:
    del wb["Service Lookup Tables"]

ws = wb.create_sheet("Service Lookup Tables", 1)

# Styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, size=13, color="FFFFFF")
subtitle_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
subtitle_font = Font(bold=True, size=11, color="FFFFFF")
col_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
col_header_font = Font(bold=True, size=10, color="000000")
category_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
note_font = Font(bold=True, size=10, color="C65911")

border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Size categories
size_categories = {
    "2-14": "Very Small Units",
    "15-30": "Small Units",
    "35-150": "Medium Units",
    "151-250": "Large Units",
    "251-400": "Very Large Units",
    "401-500": "Extra Large Units",
    "501-670": "Industrial Units",
    "671-1050": "Large Industrial",
    "1051-1500": "Very Large Industrial",
    "1501+": "Extra Large Industrial"
}

row = 1

# ============================================================================
# SERVICE A - COMPREHENSIVE INSPECTION
# ============================================================================
print("Building Service A table...")
service_a = calc_settings['serviceA']

# Service header
ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE A - COMPREHENSIVE INSPECTION"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

# Service info
ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = f"Annual Preventative Maintenance - Frequency: {service_a['frequency']} times per year (Quarterly)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

# Column headers with full descriptions
headers = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Parts & Supplies\nCost\n(inspection items)",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = col_header_font
    cell.fill = col_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[row].height = 50
row += 1

# Data rows
for kw_range, values in service_a['data'].items():
    # kW Range
    cell = ws.cell(row=row, column=1, value=f"{kw_range} kW")
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True, size=10)

    # Size Category
    cell = ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard"))
    cell.border = border
    cell.fill = category_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(size=10)

    # Labor hours
    cell = ws.cell(row=row, column=3, value=values['labor'])
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '0.0 "hrs"'

    # Mobilization hours
    cell = ws.cell(row=row, column=4, value=values['mobilization'])
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '0.0 "hrs"'

    # Parts cost
    cell = ws.cell(row=row, column=5, value=values['parts'])
    cell.border = border
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '$#,##0.00'

    # Total (formula)
    formula = f"=(C{row}+D{row})*'RFP Parameters'!$B$11+E{row}"
    cell = ws.cell(row=row, column=6, value=formula)
    cell.border = border
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '$#,##0.00'
    cell.font = Font(bold=True, size=10)

    # Frequency note
    cell = ws.cell(row=row, column=7, value="4x per year")
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

# Formula explanation
row += 1
ws.merge_cells(f'A{row}:G{row}')
cell = ws[f'A{row}']
cell.value = "Formula: Total Cost = (Labor Hours + Travel/Setup Hours) × $311.24/hr Prevailing Wage Rate + Parts & Supplies"
cell.font = Font(italic=True, size=9, color="666666")
cell.alignment = Alignment(horizontal="center")
row += 2

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 14

print("  [OK] Service A")

# ============================================================================
# SERVICE B - OIL & FILTER SERVICE
# ============================================================================
print("Building Service B table...")
service_b = calc_settings['serviceB']

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE B - OIL & FILTER SERVICE"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Oil Change & Filter Replacement - Frequency: {service_b['frequency']} time per year (Annual)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

headers_b = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Oil Filter\nReplacement\nCost",
    "Engine Oil\nVolume Needed\n(gallons)",
    "Oil Cost\n@ $24/gallon\n(with markup)",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col_idx, header in enumerate(headers_b, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = col_header_font
    cell.fill = col_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[row].height = 50
row += 1

oil_price = calc_settings['oilPrice'] * calc_settings['oilMarkup']
for kw_range, values in service_b['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).fill = category_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=3).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=5, value=values['filterCost']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=5).number_format = '$#,##0.00'

    ws.cell(row=row, column=6, value=values['oilGallons']).border = border
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=6).number_format = '0.0 "gal"'

    ws.cell(row=row, column=7, value=values['oilGallons'] * oil_price).border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    formula = f"=(C{row}+D{row})*'RFP Parameters'!$B$11+E{row}+G{row}"
    ws.cell(row=row, column=8, value=formula).border = border
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True, size=10)

    ws.cell(row=row, column=9, value="1x per year").border = border
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=9).font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

row += 1
ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Oil Pricing Details: Base cost ${calc_settings['oilPrice']:.2f}/gal × {calc_settings['oilMarkup']} markup = ${oil_price:.2f}/gal (includes overhead & handling)"
cell.font = note_font
cell.fill = note_fill
cell.alignment = Alignment(horizontal="center")
row += 2

ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 14

print("  [OK] Service B")

# ============================================================================
# SERVICE C - COOLANT SERVICE
# ============================================================================
print("Building Service C table...")
service_c = calc_settings['serviceC']

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE C - COOLANT SERVICE"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Coolant Flush & Replacement - Frequency: {service_c['frequency']} time per year (Annual/Biannual)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

headers_c = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Coolant\nVolume Needed\n(gallons)",
    "Coolant Cost\n@ $24/gallon\n(with markup)",
    "Hoses & Belts\nReplacement\nCost",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col_idx, header in enumerate(headers_c, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = col_header_font
    cell.fill = col_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[row].height = 50
row += 1

coolant_price = calc_settings['coolantPrice'] * calc_settings['coolantMarkup']
for kw_range, values in service_c['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).fill = category_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=3).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=5, value=values['coolantGallons']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=5).number_format = '0.0 "gal"'

    ws.cell(row=row, column=6, value=values['coolantGallons'] * coolant_price).border = border
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=6).number_format = '$#,##0.00'

    ws.cell(row=row, column=7, value=values['hoseBeltCost']).border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    formula = f"=(C{row}+D{row})*'RFP Parameters'!$B$11+F{row}+G{row}"
    ws.cell(row=row, column=8, value=formula).border = border
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True, size=10)

    ws.cell(row=row, column=9, value="1x per year").border = border
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=9).font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

row += 1
ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Coolant Pricing Details: Base cost ${calc_settings['coolantPrice']:.2f}/gal × {calc_settings['coolantMarkup']} markup = ${coolant_price:.2f}/gal (includes overhead & handling)"
cell.font = note_font
cell.fill = note_fill
cell.alignment = Alignment(horizontal="center")
row += 2

print("  [OK] Service C")

# ============================================================================
# SERVICE D - LAB ANALYSIS
# ============================================================================
print("Building Service D table...")
service_d = calc_settings['serviceD']

ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE D - OIL, FUEL & COOLANT ANALYSIS"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "Laboratory Testing Services - Frequency: As Needed (not kW-based)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

headers_d = [
    "Lab Analysis\nType",
    "Cost per\nSample",
    "Laboratory\nProvider",
    "Testing\nDescription",
    "When\nRequired"
]

for col_idx, header in enumerate(headers_d, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = col_header_font
    cell.fill = col_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[row].height = 40
row += 1

# Oil Analysis
ws.cell(row=row, column=1, value="Oil Analysis").border = border
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=row, column=1).font = Font(bold=True, size=10)

ws.cell(row=row, column=2, value=service_d['oilAnalysisCost']).border = border
ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=2).number_format = '$#,##0.00'
ws.cell(row=row, column=2).fill = total_fill
ws.cell(row=row, column=2).font = Font(bold=True, size=10)

ws.cell(row=row, column=3, value="External Lab").border = border
ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")

ws.cell(row=row, column=4, value="Tests for contamination, wear metals, additive levels").border = border
ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="center")

ws.cell(row=row, column=5, value="Per SOW Section 6").border = border
ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=row, column=5).font = Font(italic=True, size=9)

ws.row_dimensions[row].height = 18
row += 1

# Coolant Analysis
ws.cell(row=row, column=1, value="Coolant Analysis").border = border
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=row, column=1).font = Font(bold=True, size=10)

ws.cell(row=row, column=2, value=service_d['coolantAnalysisCost']).border = border
ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=2).number_format = '$#,##0.00'
ws.cell(row=row, column=2).fill = total_fill
ws.cell(row=row, column=2).font = Font(bold=True, size=10)

ws.cell(row=row, column=3, value="External Lab").border = border
ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")

ws.cell(row=row, column=4, value="Tests for pH, freeze point, additive depletion").border = border
ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="center")

ws.cell(row=row, column=5, value="Per SOW Section 6").border = border
ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=row, column=5).font = Font(italic=True, size=9)

ws.row_dimensions[row].height = 18
row += 1

# Fuel Analysis
ws.cell(row=row, column=1, value="Fuel Analysis").border = border
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=row, column=1).font = Font(bold=True, size=10)

ws.cell(row=row, column=2, value=service_d['fuelAnalysisCost']).border = border
ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=2).number_format = '$#,##0.00'
ws.cell(row=row, column=2).fill = total_fill
ws.cell(row=row, column=2).font = Font(bold=True, size=10)

ws.cell(row=row, column=3, value="External Lab").border = border
ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")

ws.cell(row=row, column=4, value="Tests for water content, microbes, fuel quality").border = border
ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="center")

ws.cell(row=row, column=5, value="Per SOW Section 6").border = border
ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=row, column=5).font = Font(italic=True, size=9)

ws.row_dimensions[row].height = 18
row += 1

row += 1
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "Note: Lab analysis is performed as needed, not on every generator visit. Pricing is flat per sample regardless of generator size."
cell.font = Font(italic=True, size=9, color="666666")
cell.alignment = Alignment(horizontal="center")
row += 2

ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 16

print("  [OK] Service D")

# ============================================================================
# SERVICE E - LOAD BANK TESTING
# ============================================================================
print("Building Service E table...")
service_e = calc_settings['serviceE']

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "SERVICE E - LOAD BANK TESTING"
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 25
row += 1

ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = f"Full Load Testing - Frequency: {service_e['frequency']} time per year (Annual)"
cell.font = subtitle_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 20
row += 1

headers_e = [
    "Generator\nSize Range\n(kW)",
    "Size\nCategory\nDescription",
    "Labor Time\nRequired\n(hours)",
    "Travel/Setup\nTime\n(hours)",
    "Load Bank\nEquipment\nRental",
    "Transformer\nRental\n(if needed)",
    "Delivery &\nPickup\nCost",
    "TOTAL COST\nPER VISIT\n(calculated)",
    "Service\nFrequency\nNotes"
]

for col_idx, header in enumerate(headers_e, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = col_header_font
    cell.fill = col_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[row].height = 50
row += 1

for kw_range, values in service_e['data'].items():
    ws.cell(row=row, column=1, value=f"{kw_range} kW").border = border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    ws.cell(row=row, column=2, value=size_categories.get(kw_range, "Standard")).border = border
    ws.cell(row=row, column=2).fill = category_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=3, value=values['labor']).border = border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=3).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=4, value=values['mobilization']).border = border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).number_format = '0.0 "hrs"'

    ws.cell(row=row, column=5, value=values['loadBankRental']).border = border
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=5).number_format = '$#,##0.00'

    ws.cell(row=row, column=6, value=values['transformerRental']).border = border
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=6).number_format = '$#,##0.00'

    ws.cell(row=row, column=7, value=values['deliveryCost']).border = border
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=7).number_format = '$#,##0.00'

    formula = f"=(C{row}+D{row})*'RFP Parameters'!$B$11+E{row}+F{row}+G{row}"
    ws.cell(row=row, column=8, value=formula).border = border
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True, size=10)

    ws.cell(row=row, column=9, value="1x per year").border = border
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=9).font = Font(italic=True, size=9)

    ws.row_dimensions[row].height = 18
    row += 1

row += 1
ws.merge_cells(f'A{row}:I{row}')
cell = ws[f'A{row}']
cell.value = "Note: Load bank testing runs generator at full capacity for 2 hours to verify performance under real load conditions (per SOW Section 4b)"
cell.font = Font(italic=True, size=9, color="666666")
cell.alignment = Alignment(horizontal="center")
row += 2

print("  [OK] Service E")

# Save
output_path = excel_path.replace("FINAL", "FINAL-LABELED")
print(f"\nSaving to: {output_path}")
wb.save(output_path)
wb.close()
print("[OK] Complete! File saved with fully labeled lookup tables for all 5 services.")
